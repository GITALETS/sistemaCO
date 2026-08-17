import os
import calendar
import datetime
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Meses en español
MESES_ES = [
    "", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]

# Estilos exactos para las celdas del Diagrama de Gantt
FILL_PROGRAMADO = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") # Verde claro
FONT_PROGRAMADO = Font(name="Arial", size=9, bold=True, color="000000")

FILL_REALIZADO = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Naranja / Marrón
FONT_REALIZADO = Font(name="Arial", size=9, bold=True, color="FFFFFF")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

THIN_SIDE = Side(border_style="thin", color="595959")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

GLOBAL_PROFILES = {}

def clean_text_encoding(text: str) -> str:
    """ Corrige caracteres corruptos por codificación de archivo (remueve \x00 y arregla acentos) """
    if not text:
        return ""
    clean_chars = [c for c in str(text) if (32 <= ord(c) <= 126) or (192 <= ord(c) <= 255) or c in "ÁÉÍÓÚñÑáéíóú"]
    s = "".join(clean_chars)
    s = " ".join(s.split())
    replacements = {
        "conexin": "conexión",
        "Reconexciones": "Reconexiones",
        "electrico": "eléctrico",
        "solictudes": "solicitudes",
        "daos": "daños",
        "supervicion": "supervisión",
        "Planeacion": "Planeación",
        "Atencin": "Atención",
        "SECCION": "SECCIÓN",
        "facturacion": "facturación",
        "anomalias": "anomalías",
        "conciliacion": "conciliación",
        "media  tension": "media tensión"
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.strip()

def load_job_profiles(base_dir=".") -> dict:
    """ Escanea y carga dinámicamente todos los archivos de perfil CO-03-01 *.xlsx en el directorio base """
    global GLOBAL_PROFILES
    GLOBAL_PROFILES.clear()
    profiles = {}
    if not os.path.exists(base_dir):
        return profiles

    files = [f for f in os.listdir(base_dir) if f.startswith("CO-03-01") and f.endswith(".xlsx")]
    
    for f in sorted(files):
        fpath = os.path.join(base_dir, f)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            sheet = wb["Hoja2"] if "Hoja2" in wb.sheetnames else wb.active
            
            puesto_celda = sheet["C10"].value
            puesto_name = clean_text_encoding(puesto_celda) if puesto_celda else ""
            
            if not puesto_name or puesto_name.upper() in ["TEMPORAL SINDICALIZADO", "BASE SINDICALIZADO"]:
                clean_fname = f.replace("CO-03-01 SEGUIMIENTO_PROG_ESP_TAREA", "").replace(".xlsx", "").strip()
                puesto_name = clean_text_encoding(clean_fname) if clean_fname else "AYUDANTE LINIERO (SERVICIO AL CLIENTE)"

            act_rows = [17, 19, 21, 23, 25, 27, 29, 31, 33, 35]
            activities = []
            for r in act_rows:
                val = sheet.cell(r, 2).value
                if val is not None and str(val).strip():
                    cleaned_act = clean_text_encoding(val)
                    if cleaned_act:
                        activities.append(cleaned_act)
                    
            if activities:
                key = puesto_name.upper()
                profiles[key] = {
                    "puesto": puesto_name,
                    "filename": f,
                    "activities": activities
                }
            wb.close()
        except Exception as e:
            print(f"Advertencia al leer perfil {f}: {e}")

    GLOBAL_PROFILES = profiles
    return profiles

def get_profile_activities(puesto_input: str, base_dir=".") -> list:
    """ Retorna la lista de actividades del diagrama de Gantt asociadas al puesto_input """
    if not GLOBAL_PROFILES:
        load_job_profiles(base_dir)
        
    if not puesto_input:
        return DEFAULT_ACTIVITIES
        
    target = clean_text_encoding(puesto_input).upper()
    
    # 1. Coincidencia exacta
    if target in GLOBAL_PROFILES:
        return GLOBAL_PROFILES[target]["activities"]
        
    # 2. Coincidencia parcial o por subcadena
    best_match = None
    max_score = 0
    for key, pinfo in GLOBAL_PROFILES.items():
        if key in target or target in key:
            return pinfo["activities"]
            
        target_words = set(target.split())
        key_words = set(key.split())
        common = len(target_words.intersection(key_words))
        if common > max_score:
            max_score = common
            best_match = pinfo["activities"]
            
    return best_match if best_match else DEFAULT_ACTIVITIES

def detect_worker_type(rpe: str) -> str:
    """
    Determina el tipo de trabajador según el primer carácter del RPE/RTT:
    - Inicia con LETRA ➔ TEMPORAL SINDICALIZADO (RTT)
    - Inicia con NÚMERO ➔ BASE SINDICALIZADO (RPE)
    """
    if not rpe:
        return "TEMPORAL SINDICALIZADO"
    clean_rpe = str(rpe).strip()
    if not clean_rpe:
        return "TEMPORAL SINDICALIZADO"
    if clean_rpe[0].isalpha():
        return "TEMPORAL SINDICALIZADO"
    elif clean_rpe[0].isdigit():
        return "BASE SINDICALIZADO"
    else:
        return "TEMPORAL SINDICALIZADO"

def validate_rpe(rpe: str) -> str:
    """ Valida que el RPE/RTT tenga exactamente 5 caracteres """
    clean_rpe = str(rpe).strip()
    if len(clean_rpe) != 5:
        raise ValueError(f"El RPE/RTT '{clean_rpe}' debe tener exactamente 5 caracteres (ejemplo: G982P o 84729).")
    return clean_rpe

def get_target_period(target_year: int, target_month: int):
    """ Dado un año y mes objetivo de la evaluación (ej. Julio 2026), calcula los días hábiles """
    _, total_days = calendar.monthrange(target_year, target_month)
    working_days = []
    for day in range(1, total_days + 1):
        weekday = calendar.weekday(target_year, target_month, day)
        if weekday < 5:
            working_days.append(day)
    
    month_name = MESES_ES[target_month]
    period_str = f"01 AL {total_days:02d} DE {month_name} {target_year}"
    last_day_date_str = f"{total_days:02d}/{target_month:02d}/{target_year}"
    report_date_str = f"{total_days:02d} DE {month_name} {target_year}"
    
    return {
        "year": target_year,
        "month": target_month,
        "total_days": total_days,
        "working_days": working_days,
        "period_str": period_str,
        "last_day_date_str": last_day_date_str,
        "report_date_str": report_date_str,
        "month_name": month_name
    }

def calculate_target_month_from_physical_date(physical_date_str: str):
    """ Recibe una fecha física y calcula el mes inmediato anterior para la evaluación """
    dt = None
    if isinstance(physical_date_str, datetime.date):
        dt = physical_date_str
    elif isinstance(physical_date_str, str):
        s = physical_date_str.strip()
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]:
            try:
                dt = datetime.datetime.strptime(s, fmt).date()
                break
            except ValueError:
                pass
    
    if dt is None:
        dt = datetime.date(2026, 8, 14)
        
    is_weekend = dt.weekday() >= 5
    adjusted_date = dt
    if is_weekend:
        days_to_subtract = 1 if dt.weekday() == 5 else 2
        adjusted_date = dt - datetime.timedelta(days=days_to_subtract)

    if adjusted_date.month == 1:
        target_month = 12
        target_year = adjusted_date.year - 1
    else:
        target_month = adjusted_date.month - 1
        target_year = adjusted_date.year
            
    period_info = get_target_period(target_year, target_month)
    period_info["is_weekend"] = is_weekend
    period_info["original_date"] = dt.strftime("%Y-%m-%d")
    period_info["adjusted_date"] = adjusted_date.strftime("%Y-%m-%d")
    return period_info

def generate_random_scores(min_total=80, max_total=100):
    """ Genera 10 calificaciones (5 Aptitudes, 5 Actitudes) tal que la suma total esté entre 80 y 100 """
    scores_pool = [10, 8, 6]
    while True:
        aptitudes = [random.choice(scores_pool) for _ in range(5)]
        actitudes = [random.choice(scores_pool) for _ in range(5)]
        total = sum(aptitudes) + sum(actitudes)
        if min_total <= total <= max_total:
            return aptitudes, actitudes, sum(aptitudes), sum(actitudes), total

def distribute_activities(activities_list, working_days):
    """
    Distribuye las actividades asignando EXACTAMENTE 1 bloque por cada actividad (N bloques en total).
    Garantiza:
    1. El número total de bloques marcados en la tabla es exactamente N (igual al número de actividades del perfil).
    2. Cada bloque para cada actividad tiene una duración de máximo 4 días consecutivos (entre 1 y 4 días hábiles).
    3. Los bloques se posicionan secuencialmente a lo largo del mes, evitando repetir actividades o generar marcas extra.
    """
    if not activities_list or not working_days:
        return {}
    
    n_act = len(activities_list)
    assignments = {i: [] for i in range(n_act)}
    
    # 1. Agrupar los días hábiles en secuencias de días calendarios consecutivos
    runs = []
    curr = []
    for d in working_days:
        if not curr or d == curr[-1] + 1:
            curr.append(d)
        else:
            runs.append(curr)
            curr = [d]
    if curr:
        runs.append(curr)
        
    # 2. Dividir secuencias en bloques atómicos de máximo 4 días consecutivos
    all_blocks = []
    for run in runs:
        i = 0
        while i < len(run):
            rem = len(run) - i
            if rem >= 4:
                chunk_len = 3 if rem in [5, 6] else 4
            else:
                chunk_len = rem
            all_blocks.append(run[i:i+chunk_len])
            i += chunk_len
            
    # 3. Si hay menos bloques que actividades n_act, subdividir bloques más grandes
    while len(all_blocks) < n_act:
        max_idx = -1
        max_len = 1
        for idx, b in enumerate(all_blocks):
            if len(b) > max_len:
                max_len = len(b)
                max_idx = idx
        if max_idx == -1:
            break
        b_to_split = all_blocks[max_idx]
        mid = len(b_to_split) // 2
        b1 = b_to_split[:mid]
        b2 = b_to_split[mid:]
        all_blocks = all_blocks[:max_idx] + [b1, b2] + all_blocks[max_idx+1:]

    # 4. Tomar exactamente los primeros n_act bloques (1 bloque único para cada una de las n_act actividades)
    selected_blocks = all_blocks[:n_act]
    
    for i in range(len(selected_blocks)):
        assignments[i] = sorted(selected_blocks[i])
        
    return assignments

DEFAULT_ACTIVITIES = [
    "Cortes",
    "conexión de servicio nuevo",
    "Reconexiones",
    "Restablecimiento de suministro eléctrico baja tensión",
    "Restablecimiento de suministro eléctrico media tensión",
    "Atención de solicitudes de servicio"
]

DEFAULT_OFFICIALS = {
    "responsable_seguimiento": "LIC. IVONNE REZA RUGERIO",
    "responsable_encuesta": "LIC. ERYBENALY ABARCA VARGAS",
    "jefe_area_temporal": "ING. VICENTE G. RAMOS HUERTA",
    "jefe_area_base": "ING. MARCO ANTONIO ESTRADA AMADOR",
    "jefe_area": "ING. VICENTE G. RAMOS HUERTA",
    "representante_capacitacion": "LIC. IVONNE REZA RUGERIO"
}

def fill_file_01_seguimiento(wb, data, period_info, base_dir="."):
    """ Rellena la plantilla CO-03-01 SEGUIMIENTO GANTT aplicando formatos, rellenos, bordes y actividades del perfil. """
    sheet = wb["Hoja2"] if "Hoja2" in wb.sheetnames else wb.active
    
    rpe = validate_rpe(data.get("rpe", ""))
    nombre = str(data.get("nombre", "")).strip()
    
    puesto_actual_input = str(data.get("puesto_actual", data.get("puesto_base", ""))).strip()
    puesto_actual = puesto_actual_input if puesto_actual_input else detect_worker_type(rpe)
    puesto_probar = str(data.get("puesto_probar", "AYUDANTE LINIERO (SERVICIO AL CLIENTE)")).strip()
    
    sheet["C8"] = rpe
    sheet["G8"] = nombre
    sheet["C9"] = puesto_actual
    sheet["C10"] = puesto_probar
    sheet["C11"] = "01 "
    sheet["E11"] = "AL"
    sheet["F11"] = period_info["report_date_str"]
    
    total_days = period_info["total_days"]
    has_31_days = (total_days == 31)

    for m in list(sheet.merged_cells.ranges):
        m_str = str(m)
        if "C15:" in m_str or "C40:" in m_str:
            sheet.unmerge_cells(m_str)
            
    if has_31_days:
        sheet.merge_cells("C15:AH15")
        sheet.merge_cells("C40:AH40")
        sheet.column_dimensions["AH"].width = 13.0
        
        cell_31 = sheet.cell(row=16, column=34)
        cell_31.value = 31
        cell_31.alignment = ALIGN_CENTER
        cell_31.font = Font(name="Arial", size=9, bold=True)
        cell_31.border = GRID_BORDER
    else:
        sheet.merge_cells("C15:AG15")
        sheet.merge_cells("C40:AG40")
        sheet.cell(row=16, column=34).value = None
        sheet.cell(row=16, column=34).border = Border()
        
    if "activities" in data and data["activities"]:
        activities = data["activities"]
    else:
        activities = get_profile_activities(puesto_probar, base_dir=base_dir)

    working_days = period_info["working_days"]
    assignments = distribute_activities(activities, working_days)
    
    act_rows = [17, 19, 21, 23, 25, 27, 29, 31, 33, 35]
    max_col_grid = 34 if has_31_days else 33
    
    # LIMPIEZA TOTAL PREVIA:
    # 1. Limpiar los 10 títulos de actividad en la Columna B (Filas 17, 19, 21, 23, 25, 27, 29, 31, 33, 35)
    for r in act_rows:
        sheet.cell(row=r, column=2).value = None
        
    # 2. Limpiar etiquetas P/R (Columna C) y cuadrícula del Gantt (Columnas D a AH)
    for r in range(17, 37):
        sheet.cell(row=r, column=3).value = None  # Limpiar Columna C (Etiqueta P o R)
        for c in range(4, 35):
            cell = sheet.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            if c <= max_col_grid:
                cell.border = GRID_BORDER
            else:
                cell.border = Border()

    # Rellenar únicamente las actividades del perfil activo (máximo 10)
    for idx, act_name in enumerate(activities[:10]):
        row_p = act_rows[idx]
        row_r = row_p + 1
        
        sheet.cell(row=row_p, column=2).value = clean_text_encoding(act_name)
        
        cell_label_p = sheet.cell(row=row_p, column=3)
        cell_label_p.value = "P"
        cell_label_p.alignment = ALIGN_CENTER
        
        cell_label_r = sheet.cell(row=row_r, column=3)
        cell_label_r.value = "R"
        cell_label_r.alignment = ALIGN_CENTER
        
        assigned_days = assignments.get(idx, [])
        for d in assigned_days:
            col = 3 + d
            if col <= max_col_grid:
                cell_p = sheet.cell(row=row_p, column=col)
                cell_p.value = "P"
                cell_p.fill = FILL_PROGRAMADO
                cell_p.font = FONT_PROGRAMADO
                cell_p.alignment = ALIGN_CENTER
                cell_p.border = GRID_BORDER
                
                cell_r = sheet.cell(row=row_r, column=col)
                cell_r.value = "R"
                cell_r.fill = FILL_REALIZADO
                cell_r.font = FONT_REALIZADO
                cell_r.alignment = ALIGN_CENTER
                cell_r.border = GRID_BORDER
                
    sheet["B45"] = nombre
    sheet["N45"] = data.get("responsable_seguimiento", DEFAULT_OFFICIALS["responsable_seguimiento"])
    sheet["B51"] = f"FECHA DEL REPORTE: {period_info['report_date_str']}"
    sheet["B52"] = "HORA DEL REPORTE: 16:00 HRS"

def fill_file_02_valoracion(wb, data, period_info):
    """ Rellena la plantilla CO-03-02 VALORACIÓN DE APTITUDES Y ACTITUDES """
    sheet = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
    
    rpe = validate_rpe(data.get("rpe", ""))
    nombre = str(data.get("nombre", "")).strip()
    area = str(data.get("area", "ZONA TOLUCA")).strip()
    clave = str(data.get("clave", "623X5")).strip()
    
    puesto_actual_input = str(data.get("puesto_actual", data.get("puesto_base", ""))).strip()
    puesto_actual = puesto_actual_input if puesto_actual_input else detect_worker_type(rpe)
    puesto_probar = str(data.get("puesto_probar", "AYUDANTE LINIERO (SERVICIO AL CLIENTE)")).strip()
    
    sheet["C10"] = area
    sheet["G10"] = clave
    sheet["C11"] = nombre
    sheet["G11"] = rpe
    sheet["C12"] = puesto_actual
    sheet["C13"] = puesto_probar
    sheet["G13"] = period_info["last_day_date_str"]
    
    if "aptitudes_scores" in data and "actitudes_scores" in data:
        apt_scores = data["aptitudes_scores"]
        act_scores = data["actitudes_scores"]
    else:
        apt_scores, act_scores, _, _, _ = generate_random_scores(min_total=80, max_total=100)
        
    for i, s in enumerate(apt_scores):
        sheet[f"H{19+i}"] = s
        
    for i, s in enumerate(act_scores):
        sheet[f"H{30+i}"] = s
        
    sheet["C41"] = data.get("responsable_seguimiento", DEFAULT_OFFICIALS["responsable_seguimiento"])
    sheet["B56"] = f"FECHA: {period_info['report_date_str']}"
    sheet["B57"] = "HORA: 16:00 HORAS"

def fill_file_03_encuesta(wb, data, period_info):
    """ Rellena la plantilla CO-03-03 ENCUESTA DE EVALUACIÓN """
    sheet = wb["C0-03-03"] if "C0-03-03" in wb.sheetnames else wb.active
    
    rpe = validate_rpe(data.get("rpe", ""))
    nombre = str(data.get("nombre", "")).strip()
    area = str(data.get("area", "ZONA TOLUCA")).strip()
    clave = str(data.get("clave", "DN500")).strip()
    puesto_probar = str(data.get("puesto_probar", "AYUDANTE LINIERO (SERVICIO AL CLIENTE)")).strip()
    
    sheet["D12"] = area
    sheet["D13"] = clave
    sheet["D14"] = rpe
    sheet["D15"] = puesto_probar
    sheet["D16"] = period_info["period_str"]
    
    if "survey_answers" in data:
        answers = data["survey_answers"]
        if "q2" in answers: sheet["B28"] = answers["q2"]
        if "q3" in answers: sheet["B33"] = answers["q3"]
        if "q4_why" in answers: sheet["B42"] = answers["q4_why"]
        if "q5" in answers: sheet["B46"] = answers["q5"]
        if "q6" in answers: sheet["B50"] = answers["q6"]
        if "q7" in answers: sheet["B54"] = answers["q7"]

    # Determinar si el trabajador es de Base o Temporal para asignar la firma del Jefe de Área
    worker_type = data.get("worker_type") or detect_worker_type(rpe)
    is_base = "BASE" in str(worker_type).upper() or (rpe and rpe[0].isdigit())
    
    if is_base:
        jefe_area_val = data.get("jefe_area_base") or (data.get("jefe_area") if "jefe_area" in data and data["jefe_area"] != DEFAULT_OFFICIALS["jefe_area_temporal"] else DEFAULT_OFFICIALS["jefe_area_base"])
    else:
        jefe_area_val = data.get("jefe_area_temporal") or data.get("jefe_area") or DEFAULT_OFFICIALS["jefe_area_temporal"]

    sheet["B61"] = nombre
    sheet["I61"] = data.get("responsable_encuesta", DEFAULT_OFFICIALS["responsable_encuesta"])
    sheet["B70"] = jefe_area_val
    sheet["I70"] = data.get("representante_capacitacion", DEFAULT_OFFICIALS["representante_capacitacion"])

def convert_excel_to_pdf(xlsx_path: str, pdf_path: str = None) -> str:
    """ Convierte un archivo Excel (.xlsx) a PDF (.pdf) utilizando win32com o soffice como fallback. """
    if not os.path.exists(xlsx_path):
        return None
        
    if not pdf_path:
        pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
        
    abs_xlsx = os.path.abspath(xlsx_path)
    abs_pdf = os.path.abspath(pdf_path)
    
    # 1. win32com (Excel COM automation - solo en Windows)
    if os.name == 'nt':
        try:
            import win32com.client
            pythoncom_obj = None
            try:
                import pythoncom
                pythoncom.CoInitialize()
                pythoncom_obj = pythoncom
            except Exception:
                pass
                
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            
            try:
                wb = excel.Workbooks.Open(abs_xlsx, ReadOnly=True)
                wb.ExportAsFixedFormat(0, abs_pdf)
                wb.Close(False)
            finally:
                excel.Quit()
                if pythoncom_obj:
                    try:
                        pythoncom_obj.CoUninitialize()
                    except Exception:
                        pass
            
            if os.path.exists(abs_pdf) and os.path.getsize(abs_pdf) > 0:
                return abs_pdf
        except Exception as e:
            print(f"Error al exportar a PDF vía win32com: {e}")
            
    # 2. LibreOffice soffice fallback (Linux / Windows / Mac)
    try:
        import subprocess
        import shutil
        out_dir = os.path.dirname(abs_pdf)
        soffice_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/usr/local/bin/soffice",
            "/usr/local/bin/libreoffice"
        ]
        cmd_path = None
        for sp in soffice_paths:
            if os.path.exists(sp):
                cmd_path = sp
                break
                
        if not cmd_path:
            cmd_path = shutil.which("soffice") or shutil.which("libreoffice")
            
        if cmd_path:
            cmd = [cmd_path, "--headless", "--convert-to", "pdf", abs_xlsx, "--outdir", out_dir]
            subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
            default_out = os.path.join(out_dir, os.path.splitext(os.path.basename(abs_xlsx))[0] + ".pdf")
            if os.path.exists(default_out):
                if default_out != abs_pdf:
                    os.replace(default_out, abs_pdf)
                return abs_pdf
    except Exception as e2:
        print(f"Error al exportar a PDF vía soffice: {e2}")
        
    return None

def process_single_worker(worker_data, templates_dir, output_dir, export_format="pdf"):
    """ Procesa un trabajador individual y genera los 3 archivos (PDF o Excel) en output_dir. """
    os.makedirs(output_dir, exist_ok=True)
    
    physical_date = worker_data.get("fecha_fisica", "2026-08-14")
    period_info = calculate_target_month_from_physical_date(physical_date)
    
    rpe_clean = validate_rpe(worker_data.get("rpe", ""))
    nombre_clean = str(worker_data.get("nombre", "TRABAJADOR")).strip().replace(" ", "_")
    prefix = f"{rpe_clean}_{nombre_clean}"
    
    t1_path = os.path.join(templates_dir, "CO-03-01 SEGUIMIENTO_PROG_ESP_TAREA.xlsx")
    t2_path = os.path.join(templates_dir, "CO-03-02 VALORACION_APTITUDES_ACTITUDES TAREA.xlsx")
    t3_path = os.path.join(templates_dir, "CO-03-03 ENCUESTA TAREA.xlsx")
    
    # 1. SEGUIMIENTO
    wb1 = openpyxl.load_workbook(t1_path)
    fill_file_01_seguimiento(wb1, worker_data, period_info, base_dir=templates_dir)
    out1_path = os.path.join(output_dir, f"CO-03-01_SEGUIMIENTO_{prefix}.xlsx")
    wb1.save(out1_path)
    wb1.close()
    
    # 2. VALORACIÓN
    wb2 = openpyxl.load_workbook(t2_path)
    fill_file_02_valoracion(wb2, worker_data, period_info)
    out2_path = os.path.join(output_dir, f"CO-03-02_VALORACION_{prefix}.xlsx")
    wb2.save(out2_path)
    wb2.close()
    
    # 3. ENCUESTA
    wb3 = openpyxl.load_workbook(t3_path)
    fill_file_03_encuesta(wb3, worker_data, period_info)
    out3_path = os.path.join(output_dir, f"CO-03-03_ENCUESTA_{prefix}.xlsx")
    wb3.save(out3_path)
    wb3.close()
    
    generated_xlsx_files = [out1_path, out2_path, out3_path]

    # Convertir a PDF si el formato solicitado es 'pdf'
    fmt_str = str(worker_data.get("export_format", export_format)).lower()
    if fmt_str == "pdf":
        result_files = []
        for x_path in generated_xlsx_files:
            pdf_path = os.path.splitext(x_path)[0] + ".pdf"
            converted_pdf = convert_excel_to_pdf(x_path, pdf_path)
            if converted_pdf and os.path.exists(converted_pdf) and os.path.getsize(converted_pdf) > 0:
                result_files.append(converted_pdf)
                try:
                    os.remove(x_path)
                except Exception:
                    pass
            else:
                result_files.append(x_path)
        return result_files

    return generated_xlsx_files
